"""Generates three demo files for `examples/`.

Run: `python gen_examples.py [--out examples]`

Produces:
- examples/small.csv     — 1 simple table.
- examples/medium.xlsx   — 2 sheets: a multi-row header with merged cells +
                           hidden columns + totals row, and a "vertical"
                           key/value sheet.
- examples/stress.xlsx   — 50 000 rows on a single sheet (stress test).
"""

from __future__ import annotations

import argparse
import csv
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _gen_small_csv(out: Path) -> Path:
    p = out / "small.csv"
    rows = [
        ["id", "name", "department", "salary, RUB", "hired_at", "active"],
        [1, "Alice", "Eng", 120000, "2021-03-15", "true"],
        [2, "Bob", "Eng", 135000, "2020-07-01", "true"],
        [3, "Carol", "Sales", 95000, "2022-11-20", "false"],
        [4, "Dan", "Sales", 105000, "2019-02-11", "true"],
        [5, "Eve", "HR", 88000, "2023-05-04", "true"],
    ]
    with p.open("w", encoding="utf-8", newline="") as fp:
        csv.writer(fp).writerows(rows)
    return p


def _gen_medium_xlsx(out: Path) -> Path:
    p = out / "medium.xlsx"
    wb = Workbook()
    # Sheet 1: complex header + merged cells + hidden col + totals.
    ws = wb.active
    ws.title = "Sales"

    # Row 1: top-level grouped header (merged across two-column groups).
    ws["A1"] = "Region"
    ws["B1"] = "Q1 2024"
    ws.merge_cells("B1:C1")
    ws["D1"] = "Q2 2024"
    ws.merge_cells("D1:E1")
    ws["F1"] = "Notes"

    # Row 2: sub-header.
    ws["A2"] = ""
    ws["B2"] = "Plan"
    ws["C2"] = "Fact"
    ws["D2"] = "Plan"
    ws["E2"] = "Fact"
    ws["F2"] = ""

    data = [
        ["North", 100, 92, 110, 118, "good growth"],
        ["South", 80, 88, 90, 75, "drop in Q2"],
        ["East", 120, 121, 125, 130, ""],
        ["West", 95, 90, 100, 96, ""],
        ["Итого", 395, 391, 425, 419, ""],  # totals row
    ]
    for ridx, row in enumerate(data, start=3):
        for cidx, value in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=value)

    # Hide the "Notes" column to exercise hidden-column handling.
    ws.column_dimensions["F"].hidden = True

    # Sheet 2: vertical (key/value) layout.
    ws2 = wb.create_sheet("Project")
    kv = [
        ("Project name", "Atlas"),
        ("Start date", "2024-01-15"),
        ("End date", "2024-12-31"),
        ("Budget", 1_500_000),
        ("Currency", "RUB"),
        ("Status", "active"),
    ]
    for ridx, (k, v) in enumerate(kv, start=1):
        ws2.cell(row=ridx, column=1, value=k)
        ws2.cell(row=ridx, column=2, value=v)

    # Sheet 3: tables separated by an empty row (multi-region split test).
    ws3 = wb.create_sheet("Splits")
    ws3.append(["item", "qty", "price"])
    ws3.append(["pen", 10, 1.5])
    ws3.append(["pad", 5, 4.0])
    ws3.append([])
    ws3.append(["pen", 7, 1.6])
    ws3.append(["pad", 12, 3.9])

    wb.save(p)
    return p


def _gen_stress_xlsx(out: Path, *, rows: int = 50_000) -> Path:
    p = out / "stress.xlsx"
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("events")
    ws.append(["event_id", "user_id", "ts", "category", "amount"])
    rng = random.Random(42)
    start = date(2023, 1, 1)
    categories = ["click", "view", "purchase", "signup", "logout"]
    for i in range(1, rows + 1):
        ts = start + timedelta(seconds=i * 60)
        ws.append([
            i,
            rng.randint(1, 5_000),
            ts.isoformat(),
            rng.choice(categories),
            round(rng.uniform(0, 1000), 2),
        ])
    wb.save(p)
    return p


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default="examples", type=Path)
    ap.add_argument("--stress-rows", default=50_000, type=int)
    args = ap.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)

    small = _gen_small_csv(args.out)
    medium = _gen_medium_xlsx(args.out)
    stress = _gen_stress_xlsx(args.out, rows=args.stress_rows)

    print(f"Wrote: {small}")
    print(f"Wrote: {medium}")
    print(f"Wrote: {stress} ({args.stress_rows} rows)")


if __name__ == "__main__":
    main()


# Suppress unused-import warning for openpyxl helper used in side modules.
_ = get_column_letter
