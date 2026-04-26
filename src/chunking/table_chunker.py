"""Table-aware chunking.

Chunks are formed by row ranges, never by characters. Each chunk inherits the
header from its parent region and carries the absolute sheet coordinates of
its data slice in `source_ref` so RAG answers can cite back to the source.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl.utils import get_column_letter

from src.chunking.text_projection import build_text_projection
from src.models import ChunkModel, NormalizedDocument, SourceRef, TableRegion


@dataclass(frozen=True)
class ChunkSettings:
    max_rows_per_chunk: int = 200
    max_cells_per_chunk: int = 4000
    preview_rows_in_text_projection: int = 20


def _build_a1_for_chunk(
    region_range_a1: str | None,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int,
) -> str | None:
    """Reuse the parent region's column letters when available; otherwise None
    (CSV regions don't carry meaningful A1 ranges)."""
    if region_range_a1 is None:
        return None
    return f"{get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_end}"


def build_chunks(
    doc: NormalizedDocument,
    settings: ChunkSettings | None = None,
) -> list[ChunkModel]:
    cfg = settings or ChunkSettings()
    chunks: list[ChunkModel] = []
    for sheet in doc.sheets:
        for table in sheet.table_regions:
            chunks.extend(_chunk_table(doc, table, cfg))
    return chunks


def _chunk_table(
    doc: NormalizedDocument,
    table: TableRegion,
    settings: ChunkSettings,
) -> list[ChunkModel]:
    rows = table.rows
    if not rows:
        return []

    columns_count = max(1, len(table.columns))
    max_by_cells = max(1, settings.max_cells_per_chunk // columns_count)
    rows_per_chunk = max(1, min(settings.max_rows_per_chunk, max_by_cells))

    header_context = [col.name_normalized for col in table.columns]

    # Sheet row of the first data row in this region:
    #   region.source_ref.row_start points to either header (primary region) or
    #   first data row (secondary regions, where header_rows == 0).
    data_start_in_sheet = table.source_ref.row_start + table.header_rows

    out: list[ChunkModel] = []
    for i in range(0, len(rows), rows_per_chunk):
        part = rows[i : i + rows_per_chunk]
        sheet_row_start = data_start_in_sheet + i
        sheet_row_end = sheet_row_start + len(part) - 1

        source_ref = SourceRef(
            sheet_name=table.sheet_name,
            range_a1=_build_a1_for_chunk(
                table.source_ref.range_a1,
                sheet_row_start,
                sheet_row_end,
                table.source_ref.col_start,
                table.source_ref.col_end,
            ),
            row_start=sheet_row_start,
            row_end=sheet_row_end,
            col_start=table.source_ref.col_start,
            col_end=table.source_ref.col_end,
        )
        out.append(
            ChunkModel(
                chunk_id=f"{Path(doc.source_file).stem}:{table.table_id}:{sheet_row_start}-{sheet_row_end}",
                source_file=doc.source_file,
                source_format=doc.source_format,
                sheet_name=table.sheet_name,
                table_id=table.table_id,
                row_start=sheet_row_start,
                row_end=sheet_row_end,
                source_ref=source_ref,
                header_context=header_context,
                columns=table.columns,
                records=part,
                text_projection=build_text_projection(
                    sheet_name=table.sheet_name,
                    table_id=table.table_id,
                    source_ref_a1=source_ref.range_a1,
                    header_context=header_context,
                    row_start=sheet_row_start,
                    row_end=sheet_row_end,
                    rows=part,
                    preview_rows=settings.preview_rows_in_text_projection,
                ),
            )
        )
    return out
