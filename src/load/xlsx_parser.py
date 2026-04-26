"""XLSX parser.

Reads each sheet *once* via openpyxl and:
- expands merged cells so every merged-anchor's value appears in all spanned cells,
- skips hidden rows and columns and emits a warning,
- detects a multi-row header,
- builds column names by joining values across header rows,
- preserves the original sheet coordinates so chunks can cite back to A1
  ranges in the source file.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.load.header_detection import build_column_names, detect_header_rows
from src.load.regions import build_table_regions, looks_like_vertical_grid
from src.load.type_inference import split_name_and_units
from src.models import NormalizedDocument, SheetModel


def _build_a1(row_start: int, row_end: int, col_start: int, col_end: int) -> str:
    return f"{get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_end}"


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Excel-stored dates: ISO is the safest neutral representation.
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _read_sheet(
    ws,
) -> tuple[list[list[str]], list[bool], list[bool], int, int]:
    """Return (rows_as_strings, row_hidden_flags, col_hidden_flags, max_row, max_col).

    Merged cells are expanded so each cell carries the merged-anchor's value.
    """
    max_row = max(ws.max_row or 0, 1)
    max_col = max(ws.max_column or 0, 1)

    grid: list[list[str]] = [["" for _ in range(max_col)] for _ in range(max_row)]
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False):
        for cell in row:
            grid[cell.row - 1][cell.column - 1] = _format_cell_value(cell.value)

    # Expand merged ranges: copy anchor value to every spanned cell.
    for mrange in ws.merged_cells.ranges:
        anchor = grid[mrange.min_row - 1][mrange.min_col - 1]
        if anchor == "":
            continue
        for r in range(mrange.min_row - 1, mrange.max_row):
            for c in range(mrange.min_col - 1, mrange.max_col):
                grid[r][c] = anchor

    row_hidden = [
        bool(ws.row_dimensions[r + 1].hidden) if (r + 1) in ws.row_dimensions else False
        for r in range(max_row)
    ]
    col_hidden = [
        bool(ws.column_dimensions[get_column_letter(c + 1)].hidden)
        if get_column_letter(c + 1) in ws.column_dimensions
        else False
        for c in range(max_col)
    ]
    return grid, row_hidden, col_hidden, max_row, max_col


def _filter_hidden(
    grid: list[list[str]],
    row_hidden: list[bool],
    col_hidden: list[bool],
) -> tuple[list[list[str]], list[int], list[int]]:
    """Drop hidden rows and columns. Returns (visible_grid, kept_row_indices,
    kept_col_indices). Index lists are 1-based original sheet positions, so we
    can still emit accurate `source_ref.range_a1` for the visible block."""
    kept_rows = [i for i, h in enumerate(row_hidden) if not h]
    kept_cols = [j for j, h in enumerate(col_hidden) if not h]
    if not kept_rows or not kept_cols:
        return [], [r + 1 for r in kept_rows], [c + 1 for c in kept_cols]
    visible = [[grid[r][c] for c in kept_cols] for r in kept_rows]
    return visible, [r + 1 for r in kept_rows], [c + 1 for c in kept_cols]


def parse_xlsx(file_path: str | Path, *, top_n: int = 5) -> NormalizedDocument:
    path = Path(file_path)
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheets: list[SheetModel] = []

    for ws in workbook.worksheets:
        grid, row_hidden_flags, col_hidden_flags, _max_row, _max_col = _read_sheet(ws)

        extra_warnings: list[str] = []
        if any(row_hidden_flags):
            extra_warnings.append("hidden_rows_present")
        if any(col_hidden_flags):
            extra_warnings.append("hidden_columns_present")

        visible_grid, kept_rows_1based, kept_cols_1based = _filter_hidden(
            grid, row_hidden_flags, col_hidden_flags
        )
        if not visible_grid or not kept_rows_1based or not kept_cols_1based:
            continue

        width = len(visible_grid[0])

        # Vertical (key/value) layouts: skip header detection entirely so the
        # first row's content is preserved and transposed alongside the rest.
        is_vertical = looks_like_vertical_grid(visible_grid)
        if is_vertical:
            header_rows_count = 0
            raw_names = [f"col_{i + 1}" for i in range(width)]
            norm_names = list(raw_names)
            units: list[str | None] = [None] * width
            rows_dicts = [
                {norm_names[i]: cell for i, cell in enumerate(row)} for row in visible_grid
            ]
            row_offset = kept_rows_1based[0]
        else:
            header_rows_count = detect_header_rows(visible_grid)
            header_block = visible_grid[:header_rows_count]
            data_rows_lists = visible_grid[header_rows_count:]

            name_pairs = build_column_names(header_block, width)
            raw_names = [r for r, _ in name_pairs]
            norm_names = [n for _, n in name_pairs]
            units = [split_name_and_units(r)[1] for r in raw_names]

            rows_dicts = [
                {norm_names[i]: cell for i, cell in enumerate(row)}
                for row in data_rows_lists
            ]

            # Sheet coordinates: row_offset is the absolute sheet row of the first
            # data row (1-based). Even with hidden rows skipped we keep the file's
            # original numbering — this is what users will see in Excel.
            if header_rows_count >= len(kept_rows_1based):
                # Degenerate sheet: only header, no data.
                row_offset = kept_rows_1based[-1] + 1
            else:
                row_offset = kept_rows_1based[header_rows_count]

        col_start = kept_cols_1based[0]
        col_end = kept_cols_1based[-1]

        table_id_prefix = f"{path.stem}_{ws.title}_t1"
        regions = build_table_regions(
            sheet_name=ws.title,
            table_id_prefix=table_id_prefix,
            column_names_raw=raw_names,
            column_names_normalized=norm_names,
            column_units=units,
            rows=rows_dicts,
            header_rows=header_rows_count,
            row_offset=row_offset,
            col_start=col_start,
            col_end=col_end,
            range_a1_func=_build_a1,
            extra_warnings=extra_warnings,
            top_n=top_n,
            orientation_hint="vertical" if is_vertical else None,
        )

        sheets.append(SheetModel(sheet_name=ws.title, table_regions=regions))

    workbook.close()
    return NormalizedDocument(
        source_file=str(path),
        source_format="xlsx",
        processed_at=datetime.now(timezone.utc),
        sheets=sheets,
    )
