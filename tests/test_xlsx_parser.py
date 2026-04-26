from pathlib import Path

import pytest
from openpyxl import Workbook

from src.load.xlsx_parser import parse_xlsx


@pytest.fixture()
def simple_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "name", "salary"])
    ws.append([1, "Alice", 100])
    ws.append([2, "Bob", 200])
    wb.save(p)
    return p


@pytest.fixture()
def multi_header_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "complex.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Region"
    ws["B1"] = "Q1"
    ws.merge_cells("B1:C1")
    ws["D1"] = "Q2"
    ws.merge_cells("D1:E1")
    ws["A2"] = ""
    ws["B2"] = "Plan"
    ws["C2"] = "Fact"
    ws["D2"] = "Plan"
    ws["E2"] = "Fact"
    ws.append(["North", 100, 92, 110, 118])
    ws.append(["South", 80, 88, 90, 75])
    wb.save(p)
    return p


@pytest.fixture()
def hidden_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "hidden.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "name", "secret"])
    ws.append([1, "Alice", "x"])
    ws.append([2, "Bob", "y"])
    ws.column_dimensions["C"].hidden = True
    ws.row_dimensions[3].hidden = True
    wb.save(p)
    return p


def test_simple_xlsx(simple_xlsx: Path):
    doc = parse_xlsx(simple_xlsx)
    assert len(doc.sheets) == 1
    region = doc.sheets[0].table_regions[0]
    assert region.header_rows == 1
    assert [c.name_normalized for c in region.columns] == ["id", "name", "salary"]
    assert region.row_count == 2
    assert region.source_ref.range_a1 == "A1:C3"
    assert region.source_ref.row_start == 1
    assert region.source_ref.row_end == 3


def test_multi_row_header_with_merges(multi_header_xlsx: Path):
    doc = parse_xlsx(multi_header_xlsx)
    region = doc.sheets[0].table_regions[0]
    # Header should be 2 rows, with merged "Q1" expanded to both Plan & Fact.
    assert region.header_rows == 2
    norm = [c.name_normalized for c in region.columns]
    assert norm[0] == "region"
    assert norm[1] == "q1_plan"
    assert norm[2] == "q1_fact"
    assert norm[3] == "q2_plan"
    assert norm[4] == "q2_fact"
    # Data: 2 rows starting at sheet row 3.
    assert region.row_count == 2
    assert region.source_ref.row_start == 1
    assert region.source_ref.row_end == 4


def test_hidden_rows_and_cols_are_skipped_and_warned(hidden_xlsx: Path):
    doc = parse_xlsx(hidden_xlsx)
    region = doc.sheets[0].table_regions[0]
    # Hidden col C ("secret") removed.
    assert [c.name_normalized for c in region.columns] == ["id", "name"]
    # Hidden row 3 removed -> only 1 visible data row.
    assert region.row_count == 1
    warnings = region.profile.warnings
    assert "hidden_rows_present" in warnings
    assert "hidden_columns_present" in warnings


def test_xlsx_chunking_uses_real_sheet_rows(multi_header_xlsx: Path):
    """End-to-end check that chunk row coordinates point to the actual file rows."""
    from src.chunking.table_chunker import ChunkSettings, build_chunks

    doc = parse_xlsx(multi_header_xlsx)
    chunks = build_chunks(doc, ChunkSettings(max_rows_per_chunk=1, max_cells_per_chunk=10000))
    assert len(chunks) == 2
    # First data row sits at sheet row 3 (after the 2-row header).
    assert chunks[0].row_start == 3
    assert chunks[0].row_end == 3
    assert chunks[1].row_start == 4
    assert chunks[1].row_end == 4
